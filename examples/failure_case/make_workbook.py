from openpyxl import Workbook


wb = Workbook()

ws = wb.active
# Add sample data for FILTER formula
ws["A2"] = "Apple"
ws["A3"] = "Banana"
ws["A4"] = "Cherry"
ws["A5"] = "Date"
ws["A6"] = "Elderberry"
ws["A7"] = "Fig"
ws["A8"] = "Grape"
ws["A9"] = "Honeydew"
ws["A10"] = "Kiwi"

ws["B2"] = 1
ws["B3"] = 0
ws["B4"] = 1
ws["B5"] = 0
ws["B6"] = 1
ws["B7"] = 0
ws["B8"] = 1
ws["B9"] = 0
ws["B10"] = 1

ws["D1"] = "=_xlfn.FILTER(A2:A10,B2:B10>0)"  # example dynamic array formula

# wb.calculation.fullCalcOnLoad = True  # ensure recalc on open
wb.save("out.xlsx")